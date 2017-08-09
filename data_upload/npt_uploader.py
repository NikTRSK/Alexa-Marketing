import datetime
import openpyxl
from npt_db import NPTDB
from promo_db import PromoDB
from helpers import find_and_replace_list_item_in_place

def upload_npt(db, input_file = ".\\data\\npt.xlsx"):
    workbook = openpyxl.load_workbook(input_file)
    sheet_names = workbook.get_sheet_names()
    for sheet_name in sheet_names:
        worksheet = workbook.get_sheet_by_name(sheet_name)
        # Get header of the document. Used for the table column names
        header = get_header(worksheet)
        row_count = worksheet.max_row
        column_count = worksheet.max_column
        print("Uploading NPT data for " + sheet_name)
        for row in range(11, row_count):
            item = {}
            for col in range(1, column_count):
                val = worksheet.cell(row=row, column=col).value
                if isinstance(val, datetime.datetime):
                    val = str(val)
                item[header[col-1]] = val
            if item['SegmentType'] == 'PROMO':
                item.pop('SegmentType')
                hyperlink = item['PROMO_TITLE'].split('","')
                # Extract promo title and hyperlink for video
                promo_title = hyperlink[1][:-2]
                hyperlink = hyperlink[0].replace('=HYPERLINK("', '')
                # Add in the extra fields in to the table entry
                item['PROMO_TITLE'] = promo_title
                item['VIDEO_LINK'] = hyperlink
                item['SHOW_TITLE'] = 'n/a' # n/a means the source doesn't provide the data
                item['PROMO_ID'] = 'n/a'
                item['CHANNEL_TIME_ZONE'] = 'n/a'
                item['PROJECT_ID'] = 'n/a'
                item['APPROVAL_DATE'] = 'n/a'
                item['STATUS'] = 'n/a'
                item['DIGITAL_AIR_DATE'] = 'n/a'
                item['DIGITAL_PLATFORM'] = 'n/a'
                item['SHOW_AIR_TIME'] = 'n/a'
                item['SOURCE_DB'] = 'NPT'
                db.add(item)

# Remap the header to match the database
def get_header(worksheet):
    header = []
    for i in range(1, 8):
        header.append(worksheet.cell(row=10, column=i).value)
    
    mappings = { 'Network': 'NETWORK', 'Date': 'AIR_DATE', 'DESCRIPTION': 'PROMO_TITLE',
                 'DURATION': 'PROMO_LENGTH', 'StartTime': 'AIR_TIME' }
    
    for key, value in mappings.items():
        find_and_replace_list_item_in_place(header, key, value)

    return header

# Returns a list of all the promo titles in NPT
def get_npt_titles(input_file = ".\\data\\npt.xlsx"):
    promo_list = set()
    show_list = set()
    workbook = openpyxl.load_workbook(input_file)
    sheet_names = workbook.get_sheet_names()
    for sheet_name in sheet_names:
        worksheet = workbook.get_sheet_by_name(sheet_name)
        header = get_header(worksheet)
        row_count = worksheet.max_row
        column_count = worksheet.max_column
        for row in range(11, row_count):
            item = {}
            for col in range(1, column_count):
                val = worksheet.cell(row=row, column=col).value
                if isinstance(val, datetime.datetime):
                    val = str(val)
                item[header[col-1]] = val
            segment_type = item['SegmentType']
            if segment_type in 'PROMO':
                # Add show title
                show = (item['PROMO_TITLE'].split(',')[1][1:-2]).split(':')
                if len(show) > 0:
                    show_list.add(show[0])
                # Add promo title
                promo_list.add(item['PROMO_TITLE'].split(',')[1][1:-2])
    return promo_list, show_list